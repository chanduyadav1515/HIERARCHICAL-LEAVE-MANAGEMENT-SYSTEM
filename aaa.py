from tkinter import *
from openpyxl import load_workbook
from tkinter import messagebox
from datetime import date
from tkinter import filedialog
from docx import Document
import time
import requests


today_date = date.today()   # get today's date

main = Tk()
main.geometry('650x400')
main.configure(bg="white")
main.title("SUBMISSION OF LETTER")
wb = load_workbook('index.xlsx')
sheet = wb['apple']
row_no = sheet.max_row+1  # to write data into that row

def starting():
    x = variable_dept.get()+variable_hod.get()+variable_dean.get()
    print(x)

def department


# select DEPARTMENT
variable_dept = StringVar()
OptionList_dept = ["###","EEE", "ECE", "CSE", "ITD" ] 
variable_dept.set(OptionList_dept[0])
# select HOD
variable_hod = StringVar()
OptionList_hod = ["###","EEE", "ECE", "CSE", "ITD"] 
variable_hod.set(OptionList_hod[0])
# select DEAN
variable_dean = StringVar()
OptionList_dean = ["###","ACADEMICS", "STUDENT AFFAIRS", "RESEARCH", "FACULTY"] 
variable_dean.set(OptionList_dean[0])

# options for DEPT
opt_dept = OptionMenu(main, variable_dept, *OptionList_dept)
opt_dept.config(width=5, font=('Helvetica', 12))
opt_dept.place(x=220, y=40)
opt_dept.config(bg = "white") 
opt_dept["menu"].config(bg="white")
# options for HOD
opt_hod = OptionMenu(main, variable_hod, *OptionList_hod)
opt_hod.config(width=5, font=('Helvetica', 12))
opt_hod.place(x=220, y=220)
opt_hod.config(bg = "white") 
opt_hod["menu"].config(bg="white")

# options for DEAN
opt_dean = OptionMenu(main, variable_dean, *OptionList_dean)
opt_dean.config(width=15, font=('Helvetica', 12))
opt_dean.place(x=220, y=305)
opt_dean.config(bg = "white") 
opt_dean["menu"].config(bg="white")


# LABELS ----------------------------
label_F =Label(main, text="FACULTY DETAILS ", font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white")
label_dept =Label(main, text="Select Department", font=('Helvetica', 12), fg='blue',bg="white").place(x=20, y=45)
label_Hod =Label(main, text="HEAD OF THE DEPARTMENT", font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white").place(x=20, y=185)
label_hod =Label(main, text="Select Department ", font=('Helvetica', 12), fg='blue', bg="white") .place(x=20, y=225)
label_Dean =Label(main, text="DEAN", font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white").place(x=20, y=275)
label_hod =Label(main, text="Select Department ", font=('Helvetica', 12), fg='blue', bg="white").place(x=20, y=310)
label_temp =Label(main, text="          ", font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white")
label_temp.place(x=420, y=110)

b1 = Button(main, text = "start",command = starting, font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white").place(x=70, y=45)


variable_dept.trace("w", department)
variable_cat.trace("w", category)
variable_dean.trace("w", dean)
x = "            "

main.mainloop()
