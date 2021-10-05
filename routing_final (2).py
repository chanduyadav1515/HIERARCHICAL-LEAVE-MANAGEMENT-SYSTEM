from tkinter import *
from openpyxl import load_workbook
from tkinter import messagebox
from datetime import date
from tkinter import filedialog
from docx import Document
import requests

today_date = date.today()   # get today's date
server_url = 'http://tejusolutions.com/rohit/index.php/'

main = Tk()
main.geometry('650x400')
main.configure(bg="white")
main.title("SUBMISSION OF LETTER")
wb = load_workbook('index.xlsx')
sheet = wb['apple']
row_no = sheet.max_row+1  # to write data into that row

# select DEPARTMENT
variable_dept = StringVar()
OptionList_dept = ["EEE", "ECE", "CSE", "ITD" ]
variable_dept.set(OptionList_dept[0])
# select FACULTY-EEE
variable_f = StringVar()
OptionList= ["SELECT"]
variable_f.set(OptionList[0])
# select CATEGORY
variable_cat= StringVar()
OptionList_cat = ["CL ", "CCL", "EL ", "OD " ]
variable_cat.set(OptionList_cat[0])
# select HOD
variable_hod = StringVar()
OptionList_hod = ["EEE", "ECE", "CSE", "ITD"]
variable_hod.set(OptionList_hod[0])
# select DEAN
variable_dean = StringVar()
OptionList_dean = ["ACADEMICS", "STUDENT AFFAIRS", "RESEARCH", "FACULTY"]
variable_dean.set(OptionList_dean[0])

# options for DEPT
opt_dept = OptionMenu(main, variable_dept, *OptionList_dept)
opt_dept.config(width=5, font=('Helvetica', 12))
opt_dept.place(x=220, y=40)
opt_dept.config(bg = "white")
opt_dept["menu"].config(bg="white")
# options for HOD
opt_hod = OptionMenu(main, variable_hod, *OptionList_hod)
opt_hod.config(width=5, font=('Helvetica', 12))
opt_hod.place(x=220, y=220)
opt_hod.config(bg = "white")
opt_hod["menu"].config(bg="white")
opt_hod['state'] = DISABLED
# options for DEAN
opt_dean = OptionMenu(main, variable_dean, *OptionList_dean)
opt_dean.config(width=15, font=('Helvetica', 12))
opt_dean.place(x=220, y=305)
opt_dean.config(bg = "white")
opt_dean["menu"].config(bg="white")
opt_dean['state'] = DISABLED

# LABELS ----------------------------
label_F =Label(main, text="FACULTY DETAILS ", font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white")
label_dept =Label(main, text="Select Department", font=('Helvetica', 12), fg='blue',bg="white").place(x=20, y=45)
label_Hod =Label(main, text="HEAD OF THE DEPARTMENT", font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white").place(x=20, y=185)
label_hod =Label(main, text="Select Department ", font=('Helvetica', 12), fg='blue', bg="white") .place(x=20, y=225)
label_Dean =Label(main, text="DEAN", font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white").place(x=20, y=275)
label_hod =Label(main, text="Select Department ", font=('Helvetica', 12), fg='blue', bg="white").place(x=20, y=310)
label_temp =Label(main, text="", font=('Helvetica ', 12 ,'underline'), fg='maroon', bg="white")
label_temp.place(x=420, y=110)

def department(*args):
    #labelTest.configure(text="The selected item is {}".format(variable1.get()))
    a = variable_dept.get()
    label_temp.config(text = label_temp['text']+ a +"_")
    label_faculty =Label(main, text="Select Faculty ID", font=('Helvetica', 12), fg='blue', bg="white")
    label_faculty .place(x=20, y=90)
    if (a=="EEE"): OptionList = ["01", "02", "03", "04" ]    # options for f_eee
    elif (a=="ECE"): OptionList = ["01", "02", "03", "04" ]    # options for f_ece
    elif (a=="CSE"): OptionList = ["01", "02", "03", "04" ]    # options for f_cse
    elif (a=="ITD"): OptionList = ["01", "02", "03", "04" ]    # options for f_IT
    global opt_f
    opt_f = OptionMenu(main, variable_f, *OptionList)
    opt_f.config(width=10, font=('Helvetica', 12))
    opt_f.place(x=220, y=85)
    opt_f.config(bg = "white")
    opt_f["menu"].config(bg="white")
    opt_dept['state'] = DISABLED

def faculty(*args):
    a = variable_f.get()
    label_temp.config(text = label_temp['text']+ a +"_")
    label_cat =Label(main, text="Select Category", font=('Helvetica', 12), fg='blue', bg="white")
    label_cat .place(x=20, y=140)
    global opt_cat
    opt_cat = OptionMenu(main, variable_cat, *OptionList_cat)
    opt_cat.config(width=5, font=('Helvetica', 12))
    opt_cat.place(x=220, y=135)
    opt_cat.config(bg = "white")
    opt_cat["menu"].config(bg="white")
    opt_f['state'] = DISABLED

def hod(*args):
    a = variable_hod.get()
    label_temp.config(text = label_temp['text']+ a +"_")
    opt_dean['state'] = NORMAL
    opt_hod['state'] = DISABLED
    if (a=="EEE"): col = "C"+str(row_no)
    elif (a=="ECE"): col = "D"+str(row_no)
    elif (a=="CSE"):  col = "E"+str(row_no)
    elif (a=="ITD"):   col = "F"+str(row_no)
    print(col)
    sheet[col] = 1
def category(*args):
    a = variable_cat.get()
    label_temp.config(text = label_temp['text']+ a +"_")
    opt_hod['state'] = NORMAL
    opt_cat['state'] = DISABLED

def dean(*args):
    a = variable_dean.get() #"ACADEMICS", "STUDENT AFFAIRS", "RESEARCH", "FACULTY"
    if (a=="ACADEMICS"):
        a = "ACA"
        col = "G"+str(row_no)
    elif (a=="STUDENT AFFAIRS"):
        a = "STA"
        col = "H"+str(row_no)
    elif (a=="RESEARCH"):
        a = "REA"
        col = "I"+str(row_no)
    elif (a=="FACULTY"):
        a = "FAC"
        col = "J"+str(row_no)
    label_temp.config(text = label_temp['text']+ a +"_")
    row = "A"+str(row_no)
    sheet[row] = label_temp['text'] +str(row_no)
    sheet[col] = 0
    row = "K"+str(row_no) # for principal
    sheet[row] = 0
    row = "L"+str(row_no) # for chairman
    sheet[row] = 0
    row = "b"+str(row_no) # for faculty
    sheet[row] = 9
    messagebox.showinfo("Submission"," Please Select the Letter to be submitted. " )
    wb.save('index.xlsx')
    opt_dean['state'] =  DISABLED
    chk = 1
    filename=str()
    while (chk==1):
        filename =  filedialog.askopenfilename(title = "Choose your letter file",filetypes = (("documents","*.docx"),("all files","*.*")))
        if len(filename) != 0:
            messagebox.showinfo(" SUBMISSION " , "SUBMITTED SUCCESSFULLY\n Please remember your Request ID :   " +str(row_no) )
            chk = 2
        else:
            messagebox.showinfo("Cant avoid", "Please select your word letter file ")
        

    
    document = Document(filename)
    
    document.save(label_temp['text'] +str(row_no)+".docx")
    a=label_temp['text'] +str(row_no)+".docx"
    file=open(a,'rb')
    server_url = 'http://tejusolutions.com/rohit/index.php/'
    files = {'photo':file}
    data = {'submit':True}
    response = requests.post(server_url, files=files, data=data)
        #stat["text"] = 'status : '+str(response.status_code)
    file1=open('index.xlsx','rb')
    server_url = 'http://tejusolutions.com/rohit/index.php/'
    files = {'photo':file1}
    data = {'submit':True}
    response = requests.post(server_url, files=files, data=data)
        #stat["text"] = 'status : '+str(response.status_code)
    print("done")
    main.destroy()

variable_dept.trace("w", department)
variable_f.trace("w", faculty)
variable_hod.trace("w", hod)
variable_cat.trace("w", category)
variable_dean.trace("w", dean)

main.mainloop()
