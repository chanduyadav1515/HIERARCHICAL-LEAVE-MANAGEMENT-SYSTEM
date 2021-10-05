from tkinter import *
from openpyxl import load_workbook
from tkinter import messagebox
from datetime import date
from tkinter import filedialog
from docx import Document
import time
import requests

main = Tk()
main.geometry('650x400')
main.configure(bg="white")
main.title("STATUS")
wb = load_workbook('index.xlsx')
sheet = wb['apple']
col_no = sheet.max_column
print(col_no)
def readLabelText():
    a=int(read_id.get())
    row = "B"+str(a)
    print(row)
    value=sheet[row].value
    print(value)
    '''if value==9:
        lFrom_text=Label(main,text="Submitted by Faculty", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)'''
    if value==8:
        lFrom_text=Label(main,text="Rejected By HOD             ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    if value==7:
        lFrom_text=Label(main,text="Rejected By Dean            ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    if value==6:
        lFrom_text=Label(main,text="Rejected By Principal       ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    if value==5:
        lFrom_text=Label(main,text="Rejected By Chairman        ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    if value==4:
        lFrom_text=Label(main,text="Accepted                    ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="C"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At EEE HOD's Desk                  ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="D"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At ECE HOD's Desk                   ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="E"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At CSE HOD's Desk                   ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="F" 
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At IT HOD's Desk                   ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="G"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At ACADEMICS DEAN's Desk           ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="H"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At STUDENT AFFAIRS DEAN's Desk     ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="I"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At RESEARCH DEAN's Desk            ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="J"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="AT FACULTY DEAN's Desk             ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="K"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At PRINCIPAL's Desk                ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)
    column="L"
    mat = column+str(a)
    Value = sheet[mat].value
    if Value==1 and value==9:
        l1=Label(main,text="At CHAIRMAN's Desk                 ", background = "white", fg="black", justify="left",font = "times 12").place(x=200, y=100)


read_id = StringVar()
e1 = Entry(main,textvariable=read_id).place(x = 200, y = 50)
label_id = Label(main, text = "Enter your Id",font=('Times', 16), fg='green', bg="white").place(x =50 ,y = 50)
sbmitbtn = Button(main, text = "Submit",command=readLabelText).place(x = 400, y = 45)

lFrom=Label(main, text ="STATUS :   ", background = "white", fg="blue",font = "times 16 ")
lFrom.place(x=50, y=100)

main.mainloop()


